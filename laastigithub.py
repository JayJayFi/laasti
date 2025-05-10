import tkinter as tk
from tkinter import ttk, messagebox, font
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment  # For cell coloring and formatting
from pathlib import Path
import json
import uuid  # For unique IDs for saved colors


# --- Helper Class for Tooltips ---
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)
        self.widget.bind("<ButtonPress>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        if self.tooltip_window or not self.text: return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tooltip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify='left', background="#ffffe0",
                         relief='solid', borderwidth=1, font=("tahoma", "8", "normal"), padx=4, pady=2)
        label.pack(ipadx=1)

    def hide_tooltip(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None


# --- Constants and Configuration ---
PIGMENT_DATA = {
    "P.Y.42": {"name_full": "Rautaoksidikeltainen kullankeltainen, P.Y.42", "rgb": (193, 153, 59), "intensity": 22.5},
    "P.R.101": {"name_full": "Englanninpunainen, heleä 10A P.R.101", "rgb": (177, 66, 36), "intensity": 37.5},
    "Caput Mortuum": {"name_full": "Caput Mortuum P.R.101", "rgb": (89, 45, 45), "intensity": 42.0},
    "P.Bk.11": {"name_full": "Rautaoksidimusta P.Bk.11", "rgb": (58, 44, 38), "intensity": 45.0}
}
PIGMENT_ORDER = ["P.Y.42", "P.R.101", "Caput Mortuum", "P.Bk.11"]
MORTAR_BASE_RGB = (255, 255, 255)

# --- GitHub Friendly Save/Export Directory ---
# Attempt to use a subfolder in the user's Downloads directory
APP_DATA_SUBFOLDER_NAME = "LaastiVarit_PigmentMixer"
try:
    USER_DOWNLOADS_DIR = Path.home() / "Downloads"
    DEFAULT_SAVE_DIR = USER_DOWNLOADS_DIR / APP_DATA_SUBFOLDER_NAME
except Exception:  # Fallback if Path.home() or Downloads isn't resolvable (rare)
    DEFAULT_SAVE_DIR = Path(".") / APP_DATA_SUBFOLDER_NAME

SAVED_COLORS_FILE_NAME = "saved_pigment_colors.json"

COLOR_PALETTE = {
    "window_bg": "#ECEFF1", "frame_bg": "#FFFFFF", "text_primary": "#263238",
    "text_secondary": "#546E7A", "button_bg": "#546E7A", "button_fg": "#FFFFFF",
    "slider_trough": "#CFD8DC", "accent": "#78909C", "success": "#388E3C",
    "star_favorite": "#FFD700", "swatch_border": "#B0BEC5"
}


# --- Helper Functions ---
def rgb_to_hex(rgb_tuple):
    return f"#{int(rgb_tuple[0]):02x}{int(rgb_tuple[1]):02x}{int(rgb_tuple[2]):02x}"


def hex_to_excel_rgb(hex_color):
    hex_color = hex_color.lstrip('#')
    return "FF" + hex_color.upper() if len(hex_color) == 6 else "FFFFFFFF"


def calculate_mixed_color(pigment_percentages):
    r, g, b = MORTAR_BASE_RGB
    for key in PIGMENT_ORDER:
        if key in pigment_percentages and pigment_percentages[key] > 0:
            data, perc_frac = PIGMENT_DATA[key], pigment_percentages[key] / 100.0
            alpha = min(perc_frac * data["intensity"], 1.0)
            r = r * (1 - alpha) + data["rgb"][0] * alpha
            g = g * (1 - alpha) + data["rgb"][1] * alpha
            b = b * (1 - alpha) + data["rgb"][2] * alpha
    return (max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b)))


class PigmentMixerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Weber AK Pigment Mixer (Custom Palette)")
        self.root.configure(bg=COLOR_PALETTE["window_bg"])
        self.root.minsize(720, 680)
        self.root.maxsize(850, 800)

        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("TFrame", background=COLOR_PALETTE["frame_bg"])
        self.style.configure("TLabel", background=COLOR_PALETTE["frame_bg"], foreground=COLOR_PALETTE["text_primary"],
                             padding=3)
        self.style.configure("TButton", background=COLOR_PALETTE["button_bg"], foreground=COLOR_PALETTE["button_fg"],
                             padding=(4, 3), font=('Helvetica', 9))
        self.style.map("TButton", background=[('active', COLOR_PALETTE["accent"])])
        self.style.configure("Accent.TButton", background=COLOR_PALETTE["success"],
                             foreground=COLOR_PALETTE["button_fg"], padding=(8, 5), font=('Helvetica', 10, 'bold'))
        self.style.map("Accent.TButton", background=[('active', '#2E7D32')])
        self.style.configure("Small.TButton", padding=(2, 1), font=('Helvetica', 8))
        self.style.configure("Horizontal.TScale", troughcolor=COLOR_PALETTE["slider_trough"],
                             background=COLOR_PALETTE["button_bg"])

        self.header_font = font.Font(family="Helvetica", size=11, weight="bold")
        self.label_font = font.Font(family="Helvetica", size=9)
        self.small_font = font.Font(family="Helvetica", size=8)

        self.pigment_vars = {key: tk.DoubleVar(value=0.0) for key in PIGMENT_ORDER}
        self.pigment_labels = {key: tk.StringVar(value="0,0 %") for key in PIGMENT_ORDER}

        self.saved_colors = []
        # json_save_path will be determined by _ensure_save_dir_exists and used in load/save
        self.current_save_dir = self._ensure_save_dir_exists()  # Determine actual save dir on init
        self.json_save_path = self.current_save_dir / SAVED_COLORS_FILE_NAME
        self.load_saved_colors()

        main_content_frame = ttk.Frame(root, padding=(8, 10), style="TFrame")
        main_content_frame.pack(expand=True, fill=tk.BOTH)

        preview_frame_outer = ttk.Frame(main_content_frame, padding=3, style="TFrame")
        preview_frame_outer.pack(pady=(0, 8), padx=3, fill=tk.X)
        ttk.Label(preview_frame_outer, text="Color Preview", font=self.header_font, anchor='center').pack(pady=(0, 3),
                                                                                                          fill=tk.X)
        self.color_preview = tk.Canvas(preview_frame_outer, width=250, height=180, bg=rgb_to_hex(MORTAR_BASE_RGB),
                                       highlightthickness=1, highlightbackground=COLOR_PALETTE["accent"])
        self.color_preview.pack(expand=True, fill=tk.X, pady=(0, 3))
        self.rgb_hex_label = ttk.Label(preview_frame_outer, text="RGB: (...)\nHEX: #...", font=self.small_font,
                                       anchor='center', justify=tk.CENTER)
        self.rgb_hex_label.pack(fill=tk.X)

        sliders_frame = ttk.Frame(main_content_frame, padding=8, style="TFrame")
        sliders_frame.pack(pady=8, padx=3, fill=tk.X)
        ttk.Label(sliders_frame, text="Pigment Percentages (0-10%)", font=self.header_font).grid(row=0, column=0,
                                                                                                 columnspan=3,
                                                                                                 pady=(0, 8),
                                                                                                 sticky='w')
        for i, key in enumerate(PIGMENT_ORDER):
            label_text = PIGMENT_DATA[key]['name_full'].split(',')[0]
            ttk.Label(sliders_frame, text=label_text, font=self.label_font, wraplength=170).grid(row=i + 1, column=0,
                                                                                                 sticky='w', padx=3,
                                                                                                 pady=3)
            slider = ttk.Scale(sliders_frame, from_=0, to=10, orient=tk.HORIZONTAL, variable=self.pigment_vars[key],
                               command=lambda val, k=key: self._on_slider_change(k, val), length=180)
            slider.grid(row=i + 1, column=1, sticky='ew', padx=3, pady=3)
            ttk.Label(sliders_frame, textvariable=self.pigment_labels[key], font=self.label_font, width=7).grid(
                row=i + 1, column=2, sticky='e', padx=3, pady=3)
        sliders_frame.columnconfigure(0, weight=1);
        sliders_frame.columnconfigure(1, weight=2);
        sliders_frame.columnconfigure(2, weight=0)

        actions_frame = ttk.Frame(main_content_frame, style="TFrame", padding=(0, 3))
        actions_frame.pack(fill=tk.X, pady=(5, 8), padx=3, side=tk.BOTTOM)
        save_button = ttk.Button(actions_frame, text="Save Current Color", command=self.save_current_color_action,
                                 style="Accent.TButton")
        save_button.pack(side=tk.LEFT, padx=(0, 5), expand=True, fill=tk.X)
        export_button = ttk.Button(actions_frame, text="Export Palette to Excel", command=self.export_palette_to_excel,
                                   style="Accent.TButton")
        export_button.pack(side=tk.LEFT, padx=(5, 0), expand=True, fill=tk.X)

        saved_colors_outer_frame = ttk.LabelFrame(main_content_frame, text="Saved Color Palette", padding=(8, 5))
        saved_colors_outer_frame.pack(pady=8, padx=3, fill=tk.BOTH, expand=True)
        self.saved_colors_canvas = tk.Canvas(saved_colors_outer_frame, borderwidth=0,
                                             background=COLOR_PALETTE["frame_bg"])
        self.saved_colors_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(saved_colors_outer_frame, orient=tk.VERTICAL, command=self.saved_colors_canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.saved_colors_canvas.configure(yscrollcommand=scrollbar.set)
        self.scrollable_frame = ttk.Frame(self.saved_colors_canvas, style="TFrame")
        self.saved_colors_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.scrollable_frame.bind("<Configure>", lambda e: self.saved_colors_canvas.configure(
            scrollregion=self.saved_colors_canvas.bbox("all")))
        self.saved_colors_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.saved_colors_canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.saved_colors_canvas.bind_all("<Button-5>", self._on_mousewheel)

        self.populate_saved_colors_display()
        self.update_color_preview()

    def _on_mousewheel(self, event):
        if event.num == 5 or event.delta < 0:
            self.saved_colors_canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.saved_colors_canvas.yview_scroll(-1, "units")

    def _on_slider_change(self, pigment_key, value_str):
        try:
            f_val = float(value_str)
            self.pigment_labels[pigment_key].set(f"{f_val:,.1f} %".replace('.', ','))
        except ValueError:
            self.pigment_labels[pigment_key].set("0,0 %")
        self.update_color_preview()

    def update_color_preview(self, event=None):
        current_percentages = {key: var.get() for key, var in self.pigment_vars.items()}
        mixed_rgb, hex_color = calculate_mixed_color(current_percentages), rgb_to_hex(
            calculate_mixed_color(current_percentages))
        self.color_preview.config(bg=hex_color)
        self.rgb_hex_label.config(
            text=f"RGB: ({int(mixed_rgb[0])},{int(mixed_rgb[1])},{int(mixed_rgb[2])})\nHEX: {hex_color.upper()}")

    def _ensure_save_dir_exists(self):
        """Ensures the default save directory (Downloads/APP_DATA_SUBFOLDER_NAME) exists.
           Falls back to script's current directory if Downloads is not accessible.
           Returns the Path object of the directory to be used."""
        try:
            # Check if USER_DOWNLOADS_DIR was resolved, otherwise DEFAULT_SAVE_DIR is already based on Path(".")
            if USER_DOWNLOADS_DIR.exists() and USER_DOWNLOADS_DIR.is_dir():
                target_dir = DEFAULT_SAVE_DIR  # This is USER_DOWNLOADS_DIR / APP_DATA_SUBFOLDER_NAME
            else:  # Fallback if Downloads dir itself is not found (e.g. non-standard OS)
                target_dir = Path(".") / APP_DATA_SUBFOLDER_NAME

            target_dir.mkdir(parents=True, exist_ok=True)
            return target_dir
        except Exception as e:  # Broad exception for any OS/Path issue
            # Fallback to a subfolder in the script's current directory
            fallback_dir = Path(".") / APP_DATA_SUBFOLDER_NAME
            try:
                fallback_dir.mkdir(parents=True, exist_ok=True)
                messagebox.showwarning("Directory Warning",
                                       f"Could not use Downloads folder.\nSaving data to: {fallback_dir.resolve()}\nError: {e}")
                return fallback_dir
            except Exception as e_fallback:  # If even fallback subfolder fails, use script dir directly
                messagebox.showerror("Critical Directory Error",
                                     f"Could not create any save directory.\nSaving to script directory: {Path('.').resolve()}\nError: {e_fallback}")
                return Path(".")

    def load_saved_colors(self):
        # self.current_save_dir is set in __init__ by _ensure_save_dir_exists
        # self.json_save_path is also set in __init__ based on self.current_save_dir
        try:
            if self.json_save_path.exists():
                with open(self.json_save_path, 'r') as f:
                    self.saved_colors = json.load(f)
            else:
                self.saved_colors = []  # No file found, start fresh
        except Exception as e:
            messagebox.showerror("Load Error", f"Could not load saved colors from {self.json_save_path}: {e}")
            self.saved_colors = []

    def save_colors_to_file(self):
        # self.current_save_dir and self.json_save_path are already determined
        try:
            with open(self.json_save_path, 'w') as f:
                json.dump(self.saved_colors, f, indent=4)
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save colors to {self.json_save_path}: {e}")

    def populate_saved_colors_display(self):
        for widget in self.scrollable_frame.winfo_children(): widget.destroy()
        sorted_colors = sorted(self.saved_colors, key=lambda x: (not x.get('favorite', False), x.get('timestamp', '')))
        if not sorted_colors:
            ttk.Label(self.scrollable_frame, text="No colors saved yet.", font=self.small_font).pack(pady=5)
            return
        for color_data in sorted_colors:
            color_id, recipe, hex_val, fav = color_data['id'], color_data['recipe'], color_data['hex'], color_data.get(
                'favorite', False)
            item_frame = ttk.Frame(self.scrollable_frame, padding=3, relief=tk.SOLID, borderwidth=1)
            item_frame.pack(pady=2, padx=2, fill=tk.X)
            swatch = tk.Canvas(item_frame, width=40, height=25, bg=hex_val, highlightthickness=1,
                               highlightbackground=COLOR_PALETTE["swatch_border"])
            swatch.pack(side=tk.LEFT, padx=(0, 5))
            swatch.bind("<Button-1>", lambda e, r=recipe: self.apply_saved_color_recipe(r))
            ToolTip(swatch, f"Apply: {recipe}")
            star_char, star_color = ("★", COLOR_PALETTE["star_favorite"]) if fav else (
            "☆", COLOR_PALETTE["text_secondary"])
            fav_btn = tk.Button(item_frame, text=star_char, fg=star_color, relief=tk.FLAT,
                                command=lambda cid=color_id: self.toggle_favorite_color(cid),
                                font=('Helvetica', 11, 'bold'))
            fav_btn.pack(side=tk.LEFT, padx=3)
            ToolTip(fav_btn, "Toggle Favorite")
            del_btn = ttk.Button(item_frame, text="Del", style="Small.TButton",
                                 command=lambda cid=color_id: self.delete_saved_color(cid))
            del_btn.pack(side=tk.RIGHT, padx=3)
            ToolTip(del_btn, "Delete color")
        self.root.update_idletasks()
        self.saved_colors_canvas.configure(scrollregion=self.saved_colors_canvas.bbox("all"))

    def apply_saved_color_recipe(self, recipe):
        for key in PIGMENT_ORDER:
            val = recipe.get(key, 0.0)
            self.pigment_vars[key].set(val)
            self.pigment_labels[key].set(f"{val:,.1f} %".replace('.', ','))
        self.update_color_preview()

    def save_current_color_action(self):
        percentages = {key: round(var.get(), 1) for key, var in self.pigment_vars.items()}
        if not any(p > 0 for p in percentages.values()):
            messagebox.showinfo("Save Color", "No pigments selected.");
            return
        rgb, hex_c = calculate_mixed_color(percentages), rgb_to_hex(calculate_mixed_color(percentages))
        entry = {"id": str(uuid.uuid4()), "recipe": percentages, "rgb": rgb, "hex": hex_c, "favorite": False,
                 "timestamp": datetime.now().isoformat()}
        self.saved_colors.append(entry)
        self.save_colors_to_file()
        self.populate_saved_colors_display()
        messagebox.showinfo("Color Saved", "Current color added to palette.")

    def delete_saved_color(self, color_id):
        self.saved_colors = [c for c in self.saved_colors if c['id'] != color_id]
        self.save_colors_to_file()
        self.populate_saved_colors_display()

    def toggle_favorite_color(self, color_id):
        for c in self.saved_colors:
            if c['id'] == color_id: c['favorite'] = not c.get('favorite', False); break
        self.save_colors_to_file()
        self.populate_saved_colors_display()

    def export_palette_to_excel(self):
        if not self.saved_colors:
            messagebox.showinfo("Export Empty", "No colors in palette.");
            return

        # Use the already determined current_save_dir for export
        export_dir = self.current_save_dir
        filepath = export_dir / f"custom_palette_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = openpyxl.Workbook();
        ws = wb.active;
        ws.title = "Custom Color Palette"
        ws['A1'] = "Custom Pigment Palette Export";
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row_idx = 4
        headers = ["Favorite", "Visual Color", "Saved ID (Timestamp)"] + \
                  [PIGMENT_DATA[pk]["name_full"].split(',')[0] + " (%)" for pk in PIGMENT_ORDER] + \
                  ["Test Mortar Amount (g)", "Test Water Amount (g)"] + \
                  ["Calc. " + PIGMENT_DATA[pk]["name_full"].split(',')[0] + " for Test (g)" for pk in PIGMENT_ORDER] + \
                  ["Notes"]

        for col, h_text in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=h_text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        row_idx += 1

        sorted_colors = sorted(self.saved_colors, key=lambda x: (not x.get('favorite', False), x.get('timestamp', '')))
        DEFAULT_TEST_MORTAR_G = 20.0

        pigment_percentage_cols_letters = [openpyxl.utils.get_column_letter(4 + i) for i in range(len(PIGMENT_ORDER))]
        test_mortar_col_letter = openpyxl.utils.get_column_letter(4 + len(PIGMENT_ORDER))
        calculated_pigment_cols_start_idx = 4 + len(PIGMENT_ORDER) + 2

        for c_data in sorted_colors:
            ws.cell(row=row_idx, column=1, value="★" if c_data.get('favorite') else "")
            excel_hex = hex_to_excel_rgb(c_data['hex'])
            ws.cell(row=row_idx, column=2).fill = PatternFill(start_color=excel_hex, end_color=excel_hex,
                                                              fill_type="solid")
            ws.cell(row=row_idx, column=3, value=c_data.get('timestamp', c_data['id']))

            for i, pk in enumerate(PIGMENT_ORDER):
                stored_perc_val = c_data['recipe'].get(pk, 0.0)
                excel_decimal_val = stored_perc_val / 100.0
                cell = ws.cell(row=row_idx, column=4 + i, value=excel_decimal_val)
                cell.number_format = '0.0%'

            test_mortar_cell = ws.cell(row=row_idx, column=len(PIGMENT_ORDER) + 4, value=DEFAULT_TEST_MORTAR_G)
            test_mortar_cell.number_format = '0'
            ws.cell(row=row_idx, column=len(PIGMENT_ORDER) + 5, value="")

            for i, pk in enumerate(PIGMENT_ORDER):
                perc_cell_char = pigment_percentage_cols_letters[i]
                mortar_cell_char = test_mortar_col_letter
                formula = f"={perc_cell_char}{row_idx}*{mortar_cell_char}{row_idx}"
                calc_cell = ws.cell(row=row_idx, column=calculated_pigment_cols_start_idx + i, value=formula)
                calc_cell.number_format = '0,000'

            ws.cell(row=row_idx, column=len(headers), value="")
            row_idx += 1

        for i, col_char in enumerate([openpyxl.utils.get_column_letter(j + 1) for j in range(len(headers))]):
            max_len = len(str(headers[i]))
            for r_num in range(5, row_idx):
                cell = ws.cell(row=r_num, column=i + 1)
                cell_v_str = ""
                if cell.value is not None:
                    if cell.number_format == '0.0%' and isinstance(cell.value, (int, float)):
                        display_percentage_number = cell.value * 100
                        cell_v_str = f"{display_percentage_number:,.1f}%".replace('.', ',')
                    elif isinstance(cell.value, (int, float)) and cell.number_format == '0,000':
                        cell_v_str = f"{cell.value:,.3f}".replace('.', ',')
                    else:
                        cell_v_str = str(cell.value)
                max_len = max(max_len, len(cell_v_str))
            adjusted_width = 12 if i == 1 else (max_len + 2.5)
            ws.column_dimensions[col_char].width = max(adjusted_width, 10)

        try:
            wb.save(filepath)
            messagebox.showinfo("Export Successful",
                                f"Palette exported to:\n{filepath.resolve()}")  # Show resolved path
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not save Excel to {filepath.resolve()}:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = PigmentMixerApp(root)
    root.mainloop()
